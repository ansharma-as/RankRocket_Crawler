import asyncio
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from io import BytesIO
import base64
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import seaborn as sns
from bson import ObjectId
import pandas as pd

from app.models.schemas import SEOMetrics, CrawlResult, Recommendation
from app.services.advanced_analyzer import AdvancedSEOAnalyzer


class ReportGenerator:
    def __init__(self, db):
        self.db = db
        self.advanced_analyzer = AdvancedSEOAnalyzer()
        
    async def generate_comprehensive_report(self, submission_id: str, 
                                          format_type: str = 'json') -> Dict[str, Any]:
        """Generate comprehensive SEO report"""
        
        # Get crawl result
        crawl_result = await self.db.crawl_results.find_one(
            {"url_submission_id": submission_id}
        )
        
        if not crawl_result:
            raise ValueError("Crawl result not found")
        
        # Get recommendations
        recommendations = await self.db.recommendations.find(
            {"crawl_result_id": str(crawl_result["_id"])}
        ).to_list(length=None)
        
        # Get historical data
        historical_data = await self._get_historical_data(crawl_result["url"])
        
        # Generate report sections
        report = {
            "metadata": {
                "report_id": str(ObjectId()),
                "generated_at": datetime.utcnow().isoformat(),
                "url": crawl_result["url"],
                "submission_id": submission_id,
                "report_type": "comprehensive_seo_audit"
            },
            "executive_summary": await self._generate_executive_summary(crawl_result, recommendations),
            "technical_seo": await self._generate_technical_seo_section(crawl_result),
            "content_analysis": await self._generate_content_analysis(crawl_result),
            "performance_metrics": await self._generate_performance_section(crawl_result),
            "recommendations": await self._format_recommendations(recommendations),
            "historical_trends": historical_data,
            "competitive_analysis": await self._generate_competitive_analysis(crawl_result["url"]),
            "action_plan": await self._generate_action_plan(recommendations),
            "appendix": await self._generate_appendix(crawl_result)
        }
        
        # Format output based on requested format
        if format_type == 'pdf':
            return await self._generate_pdf_report(report)
        elif format_type == 'excel':
            return await self._generate_excel_report(report)
        elif format_type == 'html':
            return await self._generate_html_report(report)
        else:
            return report
    
    async def _generate_executive_summary(self, crawl_result: Dict, 
                                        recommendations: List[Dict]) -> Dict[str, Any]:
        """Generate executive summary"""
        seo_metrics = crawl_result["seo_metrics"]
        
        # Calculate overall SEO score
        overall_score = await self._calculate_overall_seo_score(seo_metrics, recommendations)
        
        # Count issues by priority
        issue_counts = {
            "high": len([r for r in recommendations if r.get("priority") == "high"]),
            "medium": len([r for r in recommendations if r.get("priority") == "medium"]),
            "low": len([r for r in recommendations if r.get("priority") == "low"])
        }
        
        # Key findings
        key_findings = []
        
        if not seo_metrics.get("title"):
            key_findings.append("Missing page title - critical for SEO")
        
        if not seo_metrics.get("meta_description"):
            key_findings.append("Missing meta description - affects click-through rates")
        
        if seo_metrics.get("load_time", 0) > 3:
            key_findings.append(f"Slow page load time ({seo_metrics.get('load_time', 0):.2f}s)")
        
        if not seo_metrics.get("h1_tags"):
            key_findings.append("Missing H1 tag - important for content structure")
        
        return {
            "overall_score": overall_score,
            "performance_grade": self._get_performance_grade(overall_score),
            "total_issues": sum(issue_counts.values()),
            "issue_breakdown": issue_counts,
            "key_findings": key_findings[:5],  # Top 5 findings
            "primary_recommendations": [r["title"] for r in recommendations if r.get("priority") == "high"][:3],
            "crawl_date": crawl_result["crawled_at"],
            "page_insights": {
                "word_count": seo_metrics.get("content_analysis", {}).get("word_count", 0),
                "images_count": len(seo_metrics.get("images", [])),
                "links_count": len(seo_metrics.get("internal_links", [])) + len(seo_metrics.get("external_links", [])),
                "mobile_friendly": seo_metrics.get("mobile_friendly", False)
            }
        }
    
    async def _calculate_overall_seo_score(self, seo_metrics: Dict, 
                                         recommendations: List[Dict]) -> float:
        """Calculate overall SEO score"""
        score = 100.0
        
        # Deduct points for missing critical elements
        if not seo_metrics.get("title"):
            score -= 15
        if not seo_metrics.get("meta_description"):
            score -= 10
        if not seo_metrics.get("h1_tags"):
            score -= 10
        
        # Deduct points for performance issues
        load_time = seo_metrics.get("load_time", 0)
        if load_time > 3:
            score -= min(20, (load_time - 3) * 5)
        
        # Deduct points for accessibility issues
        accessibility_score = seo_metrics.get("accessibility_score", 1.0)
        score -= (1 - accessibility_score) * 15
        
        # Deduct points based on recommendation severity
        for rec in recommendations:
            impact = rec.get("impact_score", 0.5)
            if rec.get("priority") == "high":
                score -= impact * 10
            elif rec.get("priority") == "medium":
                score -= impact * 5
            else:
                score -= impact * 2
        
        return max(0, min(100, round(score, 1)))
    
    def _get_performance_grade(self, score: float) -> str:
        """Convert score to letter grade"""
        if score >= 90:
            return "A"
        elif score >= 80:
            return "B"
        elif score >= 70:
            return "C"
        elif score >= 60:
            return "D"
        else:
            return "F"
    
    async def _generate_technical_seo_section(self, crawl_result: Dict) -> Dict[str, Any]:
        """Generate technical SEO analysis"""
        seo_metrics = crawl_result["seo_metrics"]
        
        return {
            "meta_tags": {
                "title": {
                    "content": seo_metrics.get("title"),
                    "length": len(seo_metrics.get("title", "")),
                    "status": "good" if seo_metrics.get("title") and 50 <= len(seo_metrics.get("title", "")) <= 60 else "needs_improvement"
                },
                "description": {
                    "content": seo_metrics.get("meta_description"),
                    "length": len(seo_metrics.get("meta_description", "")),
                    "status": "good" if seo_metrics.get("meta_description") and 150 <= len(seo_metrics.get("meta_description", "")) <= 160 else "needs_improvement"
                },
                "keywords": seo_metrics.get("meta_keywords"),
                "canonical": seo_metrics.get("canonical_url")
            },
            "heading_structure": {
                "h1_count": len(seo_metrics.get("h1_tags", [])),
                "h2_count": len(seo_metrics.get("h2_tags", [])),
                "h3_count": len(seo_metrics.get("h3_tags", [])),
                "hierarchy_status": "good" if len(seo_metrics.get("h1_tags", [])) == 1 else "needs_improvement"
            },
            "structured_data": {
                "schemas_found": len(seo_metrics.get("structured_data", [])),
                "types": [schema.get("type") for schema in seo_metrics.get("structured_data", [])],
                "status": "good" if seo_metrics.get("structured_data") else "missing"
            },
            "open_graph": {
                "tags_found": len(seo_metrics.get("og_tags", {})),
                "complete": len(seo_metrics.get("og_tags", {})) >= 4,
                "tags": seo_metrics.get("og_tags", {})
            },
            "twitter_cards": {
                "configured": bool(seo_metrics.get("twitter_cards", {})),
                "tags": seo_metrics.get("twitter_cards", {})
            },
            "ssl_certificate": seo_metrics.get("ssl_info", {}),
            "robots_txt": seo_metrics.get("robots_txt_info", {}),
            "sitemap": seo_metrics.get("sitemap_info", {})
        }
    
    async def _generate_content_analysis(self, crawl_result: Dict) -> Dict[str, Any]:
        """Generate content analysis section"""
        seo_metrics = crawl_result["seo_metrics"]
        content_analysis = seo_metrics.get("content_analysis", {})
        
        return {
            "content_quality": {
                "word_count": content_analysis.get("word_count", 0),
                "readability_score": content_analysis.get("readability_score", {}),
                "keyword_density": content_analysis.get("keyword_density", {}),
                "content_uniqueness": content_analysis.get("content_uniqueness", {}),
                "topic_coverage": content_analysis.get("topic_coverage", {})
            },
            "images": {
                "total_images": len(seo_metrics.get("images", [])),
                "images_with_alt": len([img for img in seo_metrics.get("images", []) if img.get("alt")]),
                "optimization_score": self._calculate_image_optimization_score(seo_metrics.get("images", []))
            },
            "links": {
                "internal_links": len(seo_metrics.get("internal_links", [])),
                "external_links": len(seo_metrics.get("external_links", [])),
                "link_ratio": len(seo_metrics.get("internal_links", [])) / max(1, len(seo_metrics.get("external_links", [])))
            },
            "content_gaps": content_analysis.get("content_gaps", []),
            "content_freshness": content_analysis.get("content_freshness", {})
        }
    
    def _calculate_image_optimization_score(self, images: List[Dict]) -> float:
        """Calculate image optimization score"""
        if not images:
            return 100.0
        
        images_with_alt = len([img for img in images if img.get("alt")])
        return (images_with_alt / len(images)) * 100
    
    async def _generate_performance_section(self, crawl_result: Dict) -> Dict[str, Any]:
        """Generate performance analysis section"""
        seo_metrics = crawl_result["seo_metrics"]
        
        return {
            "page_speed": {
                "load_time": seo_metrics.get("load_time", 0),
                "page_size": seo_metrics.get("page_size", 0),
                "status_code": seo_metrics.get("status_code", 200),
                "performance_score": self._calculate_performance_score(seo_metrics)
            },
            "core_web_vitals": seo_metrics.get("core_web_vitals", {}),
            "mobile_performance": {
                "mobile_friendly": seo_metrics.get("mobile_friendly", False),
                "responsive_design": self._check_responsive_design(seo_metrics)
            },
            "accessibility": {
                "score": seo_metrics.get("accessibility_score", 0),
                "issues": self._identify_accessibility_issues(seo_metrics)
            }
        }
    
    def _calculate_performance_score(self, seo_metrics: Dict) -> float:
        """Calculate performance score"""
        load_time = seo_metrics.get("load_time", 0)
        page_size = seo_metrics.get("page_size", 0)
        
        # Score based on load time
        time_score = max(0, 100 - (load_time * 20))
        
        # Score based on page size (penalty for large pages)
        size_score = max(0, 100 - (page_size / 50000))
        
        return round((time_score + size_score) / 2, 1)
    
    def _check_responsive_design(self, seo_metrics: Dict) -> bool:
        """Check if responsive design is implemented"""
        # Simple check based on viewport meta tag
        return seo_metrics.get("mobile_friendly", False)
    
    def _identify_accessibility_issues(self, seo_metrics: Dict) -> List[str]:
        """Identify accessibility issues"""
        issues = []
        
        # Check for images without alt text
        images = seo_metrics.get("images", [])
        images_without_alt = [img for img in images if not img.get("alt")]
        if images_without_alt:
            issues.append(f"{len(images_without_alt)} images missing alt text")
        
        # Check for proper heading hierarchy
        h1_count = len(seo_metrics.get("h1_tags", []))
        if h1_count == 0:
            issues.append("Missing H1 tag")
        elif h1_count > 1:
            issues.append("Multiple H1 tags found")
        
        return issues
    
    async def _format_recommendations(self, recommendations: List[Dict]) -> Dict[str, Any]:
        """Format recommendations by priority"""
        formatted_recs = {
            "high_priority": [],
            "medium_priority": [],
            "low_priority": [],
            "total_count": len(recommendations)
        }
        
        for rec in recommendations:
            priority = rec.get("priority", "medium")
            formatted_rec = {
                "title": rec.get("title"),
                "description": rec.get("description"),
                "current_value": rec.get("current_value"),
                "suggested_value": rec.get("suggested_value"),
                "impact_score": rec.get("impact_score", 0),
                "type": rec.get("type")
            }
            
            if priority == "high":
                formatted_recs["high_priority"].append(formatted_rec)
            elif priority == "medium":
                formatted_recs["medium_priority"].append(formatted_rec)
            else:
                formatted_recs["low_priority"].append(formatted_rec)
        
        return formatted_recs
    
    async def _get_historical_data(self, url: str) -> Dict[str, Any]:
        """Get historical crawl data for trending"""
        historical_crawls = await self.db.crawl_results.find(
            {"url": url}
        ).sort("crawled_at", -1).limit(30).to_list(length=None)
        
        if len(historical_crawls) < 2:
            return {"message": "Insufficient historical data", "crawls": []}
        
        # Extract trends
        trends = {
            "load_time_trend": [],
            "seo_score_trend": [],
            "recommendations_trend": [],
            "crawl_dates": []
        }
        
        for crawl in reversed(historical_crawls):
            trends["crawl_dates"].append(crawl["crawled_at"])
            trends["load_time_trend"].append(crawl["seo_metrics"].get("load_time", 0))
            
            # Calculate SEO score for this crawl
            recommendations = await self.db.recommendations.find(
                {"crawl_result_id": str(crawl["_id"])}
            ).to_list(length=None)
            
            seo_score = await self._calculate_overall_seo_score(crawl["seo_metrics"], recommendations)
            trends["seo_score_trend"].append(seo_score)
            trends["recommendations_trend"].append(len(recommendations))
        
        return {
            "total_crawls": len(historical_crawls),
            "trends": trends,
            "improvements": self._calculate_improvements(trends)
        }
    
    def _calculate_improvements(self, trends: Dict) -> Dict[str, Any]:
        """Calculate improvements over time"""
        improvements = {}
        
        if len(trends["seo_score_trend"]) >= 2:
            latest_score = trends["seo_score_trend"][-1]
            previous_score = trends["seo_score_trend"][-2]
            improvements["seo_score_change"] = latest_score - previous_score
        
        if len(trends["load_time_trend"]) >= 2:
            latest_load_time = trends["load_time_trend"][-1]
            previous_load_time = trends["load_time_trend"][-2]
            improvements["load_time_change"] = latest_load_time - previous_load_time
        
        return improvements
    
    async def _generate_competitive_analysis(self, url: str) -> Dict[str, Any]:
        """Generate competitive analysis (placeholder)"""
        return {
            "message": "Competitive analysis requires competitor URLs",
            "suggestions": [
                "Add competitor URLs for comparison",
                "Monitor competitor SEO performance",
                "Track ranking changes over time"
            ]
        }
    
    async def _generate_action_plan(self, recommendations: List[Dict]) -> Dict[str, Any]:
        """Generate actionable plan"""
        high_priority = [r for r in recommendations if r.get("priority") == "high"]
        medium_priority = [r for r in recommendations if r.get("priority") == "medium"]
        
        action_plan = {
            "immediate_actions": [],
            "short_term_actions": [],
            "long_term_actions": [],
            "estimated_timeline": "2-4 weeks"
        }
        
        # Immediate actions (high priority)
        for rec in high_priority[:3]:
            action_plan["immediate_actions"].append({
                "task": rec.get("title"),
                "priority": "high",
                "estimated_effort": "1-2 hours",
                "impact": rec.get("impact_score", 0.5)
            })
        
        # Short-term actions (remaining high + medium priority)
        for rec in (high_priority[3:] + medium_priority[:3]):
            action_plan["short_term_actions"].append({
                "task": rec.get("title"),
                "priority": rec.get("priority"),
                "estimated_effort": "2-4 hours",
                "impact": rec.get("impact_score", 0.5)
            })
        
        # Long-term actions
        action_plan["long_term_actions"] = [
            "Monitor SEO performance monthly",
            "Update content regularly",
            "Build quality backlinks",
            "Optimize for Core Web Vitals"
        ]
        
        return action_plan
    
    async def _generate_appendix(self, crawl_result: Dict) -> Dict[str, Any]:
        """Generate appendix with technical details"""
        return {
            "crawl_metadata": {
                "crawl_id": str(crawl_result["_id"]),
                "crawl_date": crawl_result["crawled_at"],
                "user_agent": "RankRocket/1.0",
                "crawl_method": "automated"
            },
            "technical_details": {
                "response_headers": crawl_result.get("response_headers", {}),
                "server_info": crawl_result.get("server_info", {}),
                "dns_info": crawl_result.get("dns_info", {})
            },
            "glossary": {
                "SEO": "Search Engine Optimization",
                "Core Web Vitals": "Google's metrics for page experience",
                "LCP": "Largest Contentful Paint",
                "FID": "First Input Delay",
                "CLS": "Cumulative Layout Shift"
            }
        }
    
    async def _generate_pdf_report(self, report: Dict) -> Dict[str, Any]:
        """Generate PDF report"""
        # This is a placeholder - would require HTML template and PDF generation
        return {
            "format": "pdf",
            "message": "PDF generation requires HTML template setup",
            "data": report
        }
    
    async def _generate_excel_report(self, report: Dict) -> Dict[str, Any]:
        """Generate Excel report"""
        # Create workbook
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        # Add summary data
        summary = report["executive_summary"]
        ws_summary["A1"] = "SEO Audit Report"
        ws_summary["A2"] = f"URL: {report['metadata']['url']}"
        ws_summary["A3"] = f"Generated: {report['metadata']['generated_at']}"
        ws_summary["A5"] = "Overall Score"
        ws_summary["B5"] = summary["overall_score"]
        
        # Recommendations sheet
        ws_recs = wb.create_sheet("Recommendations")
        ws_recs["A1"] = "Priority"
        ws_recs["B1"] = "Title"
        ws_recs["C1"] = "Description"
        ws_recs["D1"] = "Impact"
        
        row = 2
        for priority in ["high_priority", "medium_priority", "low_priority"]:
            for rec in report["recommendations"][priority]:
                ws_recs[f"A{row}"] = priority.replace("_priority", "")
                ws_recs[f"B{row}"] = rec["title"]
                ws_recs[f"C{row}"] = rec["description"]
                ws_recs[f"D{row}"] = rec["impact_score"]
                row += 1
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "format": "excel",
            "content": base64.b64encode(output.read()).decode(),
            "filename": f"seo_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    
    async def _generate_html_report(self, report: Dict) -> Dict[str, Any]:
        """Generate HTML report"""
        # Basic HTML template
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>SEO Audit Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .header { background-color: #f0f0f0; padding: 20px; }
                .section { margin: 20px 0; }
                .score { font-size: 24px; font-weight: bold; }
                .high-priority { color: #d32f2f; }
                .medium-priority { color: #f57c00; }
                .low-priority { color: #388e3c; }
                table { width: 100%; border-collapse: collapse; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>SEO Audit Report</h1>
                <p>URL: {url}</p>
                <p>Generated: {generated_at}</p>
                <p class="score">Overall Score: {overall_score}/100 ({grade})</p>
            </div>
            
            <div class="section">
                <h2>Executive Summary</h2>
                <p>Total Issues: {total_issues}</p>
                <ul>
                    <li>High Priority: {high_priority_count}</li>
                    <li>Medium Priority: {medium_priority_count}</li>
                    <li>Low Priority: {low_priority_count}</li>
                </ul>
            </div>
            
            <div class="section">
                <h2>Key Recommendations</h2>
                <table>
                    <tr><th>Priority</th><th>Issue</th><th>Impact</th></tr>
                    {recommendations_table}
                </table>
            </div>
        </body>
        </html>
        """
        
        # Format the template
        summary = report["executive_summary"]
        recommendations_html = ""
        
        for priority in ["high_priority", "medium_priority", "low_priority"]:
            for rec in report["recommendations"][priority]:
                priority_class = priority.replace("_priority", "-priority")
                recommendations_html += f"""
                <tr class="{priority_class}">
                    <td>{priority.replace('_priority', '').title()}</td>
                    <td>{rec['title']}</td>
                    <td>{rec['impact_score']}</td>
                </tr>
                """
        
        html_content = html_template.format(
            url=report['metadata']['url'],
            generated_at=report['metadata']['generated_at'],
            overall_score=summary['overall_score'],
            grade=summary['performance_grade'],
            total_issues=summary['total_issues'],
            high_priority_count=summary['issue_breakdown']['high'],
            medium_priority_count=summary['issue_breakdown']['medium'],
            low_priority_count=summary['issue_breakdown']['low'],
            recommendations_table=recommendations_html
        )
        
        return {
            "format": "html",
            "content": html_content,
            "filename": f"seo_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        }